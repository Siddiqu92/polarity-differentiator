"""Extract real family office data from FO-MAX Excel."""
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

from src.paths import EXCEL_FILE, EXTRACTED_CSV, BUILD_LOG, OUTPUT_DIR, LOGS_DIR

MAX_RECORDS = 50


def _cell(value, default=""):
    if value is None:
        return default
    return str(value).strip()


def extract_from_excel():
    """Extract family offices from FO-MAX-data-sample Excel (header row 4, data row 5+)."""
    if not EXCEL_FILE.exists():
        raise FileNotFoundError(f"Excel file not found: {EXCEL_FILE}")

    wb = load_workbook(EXCEL_FILE, read_only=True, data_only=True)
    ws = wb.active
    records = []

    for row in ws.iter_rows(min_row=5, max_row=120, values_only=True):
        if not row or len(row) < 2 or not row[1]:
            continue

        record = {
            "name": _cell(row[1]),
            "validation_period": _cell(row[2]),
            "data_completion_text": _cell(row[3]),
            "data_completion_visual": _cell(row[4]),
            "description": _cell(row[5] if len(row) > 5 else ""),
            "investment_thesis": _cell(row[6] if len(row) > 6 else ""),
            "sectors": _cell(row[7] if len(row) > 7 else ""),
            "domain": _cell(row[8] if len(row) > 8 else ""),
            "website": _cell(row[9] if len(row) > 9 else ""),
            "url_quality": _cell(row[10] if len(row) > 10 else ""),
            "corporate_linkedin": _cell(row[11] if len(row) > 11 else ""),
            "street_address": _cell(row[12] if len(row) > 12 else ""),
            "city": _cell(row[13] if len(row) > 13 else ""),
            "state_region": _cell(row[14] if len(row) > 14 else ""),
            "country": _cell(row[15] if len(row) > 15 else ""),
            "contact_first_name": _cell(row[16] if len(row) > 16 else ""),
            "contact_last_name": _cell(row[17] if len(row) > 17 else ""),
            "contact_full_name": _cell(row[18] if len(row) > 18 else ""),
            "contact_job_title": _cell(row[19] if len(row) > 19 else ""),
            "contact_location": _cell(row[20] if len(row) > 20 else ""),
            "contact_linkedin": _cell(row[21] if len(row) > 21 else ""),
            "contact_email": _cell(row[22] if len(row) > 22 else ""),
            "email_validation": _cell(row[23] if len(row) > 23 else ""),
            "email_validation_explanation": _cell(row[24] if len(row) > 24 else ""),
            "email_quality": _cell(row[25] if len(row) > 25 else ""),
            "contact_phone": _cell(row[26] if len(row) > 26 else ""),
            "secondary_email": _cell(row[27] if len(row) > 27 else ""),
            "secondary_email_validation": _cell(row[28] if len(row) > 28 else ""),
            "secondary_email_validation_explanation": _cell(row[29] if len(row) > 29 else ""),
            "secondary_email_quality": _cell(row[30] if len(row) > 30 else ""),
            "secondary_phone": _cell(row[31] if len(row) > 31 else ""),
        }
        records.append(record)

    wb.close()
    total = len(records)
    selected = records[:MAX_RECORDS]
    print(f"[OK] Extracted {total} family offices; returning top {len(selected)} (limit={MAX_RECORDS})")
    return selected


def save_extracted_data(records):
    """Save extracted records to CSV."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(records)
    df.to_csv(EXTRACTED_CSV, index=False, encoding="utf-8")
    print(f"[OK] Saved {len(records)} records to: {EXTRACTED_CSV}")
    return EXTRACTED_CSV


def log_extraction():
    """Log extraction process."""
    records = extract_from_excel()
    output_path = save_extracted_data(records)

    LOGS_DIR.mkdir(parents=True, exist_ok=True)
    log_entry = f"\n[DATA EXTRACTION - {datetime.now().isoformat()}]\n"
    log_entry += f"Source: {EXCEL_FILE.name}\n"
    log_entry += f"Records extracted: {len(records)}\n"
    log_entry += f"Output file: {output_path}\n"
    log_entry += f"Data quality: REAL family office records from FO-MAX verified database\n"

    with open(BUILD_LOG, "a", encoding="utf-8") as f:
        f.write(log_entry)

    print("\n[OK] Data extraction complete!")
    return records


if __name__ == "__main__":
    log_extraction()
